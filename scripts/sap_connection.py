"""sap_connection.py

Helper de conexión a SAP Business One.

Soporta 4 modos (definidos por SAP_MODE en config/.env):

    service_layer   — API REST oficial (recomendada)
    hana            — conexión directa HANA (hdbcli)
    mssql           — conexión directa SQL Server (pyodbc)
    csv_drop        — lee CSV que TI deja en una carpeta compartida

Cada backend expone la misma interfaz:

    conn = get_connection()
    rows = conn.fetch_table("OITM", since=None)   # iterable de dicts
    total = conn.count_table("OITM")              # int
    conn.close()

Uso directo para probar conexión:

    python scripts/sap_connection.py --test
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import Iterable, Iterator, Optional

ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = ROOT / "config"

try:
    from dotenv import load_dotenv
except ImportError:
    print("Falta dependencia: pip install python-dotenv", file=sys.stderr)
    sys.exit(2)

load_dotenv(CONFIG_DIR / ".env")


def _env(key: str, default: Optional[str] = None) -> Optional[str]:
    val = os.environ.get(key, default)
    if val is None or val == "":
        return default
    return val


# ---------------------------------------------------------------------------
# Interfaz común
# ---------------------------------------------------------------------------

class SAPConnection:
    """Interfaz abstracta."""

    mode: str = "abstract"

    def fetch_table(self, table: str, since: Optional[str] = None,
                    since_col: Optional[str] = None) -> Iterator[dict]:
        raise NotImplementedError

    def count_table(self, table: str) -> int:
        raise NotImplementedError

    def close(self) -> None:
        pass

    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.close()


# ---------------------------------------------------------------------------
# Service Layer (REST)
# ---------------------------------------------------------------------------

class ServiceLayerConnection(SAPConnection):
    mode = "service_layer"

    def __init__(self) -> None:
        import requests
        import urllib3

        self._requests = requests
        self.base = _env("SAP_SL_URL", "").rstrip("/")
        self.company_db = _env("SAP_SL_COMPANY_DB")
        self.user = _env("SAP_SL_USER")
        self.password = _env("SAP_SL_PASSWORD")
        self.verify_ssl = (_env("SAP_SL_VERIFY_SSL", "false") or "false").lower() == "true"
        if not self.verify_ssl:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        missing = [k for k, v in {
            "SAP_SL_URL": self.base,
            "SAP_SL_COMPANY_DB": self.company_db,
            "SAP_SL_USER": self.user,
            "SAP_SL_PASSWORD": self.password,
        }.items() if not v]
        if missing:
            raise RuntimeError(
                "Service Layer mal configurado, faltan: " + ", ".join(missing)
            )

        self.session = requests.Session()
        self.session.verify = self.verify_ssl
        self._login()

    def _login(self) -> None:
        r = self.session.post(
            f"{self.base}/Login",
            json={"CompanyDB": self.company_db, "UserName": self.user,
                  "Password": self.password},
            timeout=30,
        )
        r.raise_for_status()

    def _entity_for(self, table: str) -> str:
        """Mapea nombre de tabla SAP a EntitySet Service Layer."""
        # Importamos aquí para evitar dependencia si no se usa.
        import yaml
        with open(CONFIG_DIR / "tables.yaml", "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        for group, tables in data.items():
            if table in tables:
                entity = tables[table].get("sl_name") or table
                return entity
        return table

    def fetch_table(self, table: str, since: Optional[str] = None,
                    since_col: Optional[str] = None) -> Iterator[dict]:
        entity = self._entity_for(table)
        page_size = int(_env("EXTRACT_PAGE_SIZE", "1000") or "1000")
        url = f"{self.base}/{entity}?$top={page_size}"
        if since and since_col:
            url += f"&$filter={since_col} ge '{since}'"
        skip = 0
        while True:
            full_url = url + (f"&$skip={skip}" if skip else "")
            r = self.session.get(full_url, headers={
                "Prefer": "odata.maxpagesize=" + str(page_size)
            }, timeout=120)
            r.raise_for_status()
            payload = r.json()
            batch = payload.get("value", [])
            for row in batch:
                yield row
            if "odata.nextLink" not in payload and len(batch) < page_size:
                break
            skip += len(batch)

    def count_table(self, table: str) -> int:
        entity = self._entity_for(table)
        r = self.session.get(f"{self.base}/{entity}/$count", timeout=60)
        r.raise_for_status()
        try:
            return int(r.text.strip())
        except ValueError:
            return -1

    def close(self) -> None:
        try:
            self.session.post(f"{self.base}/Logout", timeout=10)
        except Exception:
            pass
        self.session.close()


# ---------------------------------------------------------------------------
# HANA directo
# ---------------------------------------------------------------------------

class HanaConnection(SAPConnection):
    mode = "hana"

    def __init__(self) -> None:
        try:
            from hdbcli import dbapi
        except ImportError as e:
            raise RuntimeError("Falta dependencia hdbcli: pip install hdbcli") from e

        self.schema = _env("SAP_HANA_SCHEMA")
        self.conn = dbapi.connect(
            address=_env("SAP_HANA_HOST"),
            port=int(_env("SAP_HANA_PORT", "30015") or "30015"),
            user=_env("SAP_HANA_USER"),
            password=_env("SAP_HANA_PASSWORD"),
        )

    def _select(self, table: str, cols: str = "*",
                where: Optional[str] = None) -> Iterable[tuple]:
        q = f'SELECT {cols} FROM "{self.schema}"."{table}"'
        if where:
            q += f" WHERE {where}"
        cur = self.conn.cursor()
        cur.execute(q)
        names = [d[0] for d in cur.description]
        for row in cur:
            yield dict(zip(names, row))

    def fetch_table(self, table: str, since: Optional[str] = None,
                    since_col: Optional[str] = None) -> Iterator[dict]:
        where = None
        if since and since_col:
            where = f'"{since_col}" >= DATE \'{since}\''
        yield from self._select(table, where=where)

    def count_table(self, table: str) -> int:
        cur = self.conn.cursor()
        cur.execute(f'SELECT COUNT(*) FROM "{self.schema}"."{table}"')
        return int(cur.fetchone()[0])

    def close(self) -> None:
        try:
            self.conn.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# MSSQL directo
# ---------------------------------------------------------------------------

class MssqlConnection(SAPConnection):
    mode = "mssql"

    def __init__(self) -> None:
        try:
            import pyodbc
        except ImportError as e:
            raise RuntimeError("Falta dependencia pyodbc: pip install pyodbc") from e

        driver = _env("SAP_MSSQL_DRIVER", "ODBC Driver 18 for SQL Server")
        conn_str = (
            f"DRIVER={{{driver}}};"
            f"SERVER={_env('SAP_MSSQL_HOST')},{_env('SAP_MSSQL_PORT','1433')};"
            f"DATABASE={_env('SAP_MSSQL_DB')};"
            f"UID={_env('SAP_MSSQL_USER')};"
            f"PWD={_env('SAP_MSSQL_PASSWORD')};"
            "TrustServerCertificate=yes;"
        )
        self.conn = pyodbc.connect(conn_str, timeout=30)

    def fetch_table(self, table: str, since: Optional[str] = None,
                    since_col: Optional[str] = None) -> Iterator[dict]:
        q = f"SELECT * FROM [{table}]"
        params: tuple = ()
        if since and since_col:
            q += f" WHERE [{since_col}] >= ?"
            params = (since,)
        cur = self.conn.cursor()
        cur.execute(q, *params)
        cols = [c[0] for c in cur.description]
        for row in cur:
            yield dict(zip(cols, row))

    def count_table(self, table: str) -> int:
        cur = self.conn.cursor()
        cur.execute(f"SELECT COUNT(*) FROM [{table}]")
        return int(cur.fetchone()[0])

    def close(self) -> None:
        try:
            self.conn.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# CSV drop (fallback)
# ---------------------------------------------------------------------------

class CsvDropConnection(SAPConnection):
    mode = "csv_drop"

    def __init__(self) -> None:
        import csv as _csv
        self._csv = _csv
        self.drop_dir = Path(_env("SAP_CSV_DROP_DIR", "") or "")
        if not self.drop_dir.is_dir():
            raise RuntimeError(
                f"SAP_CSV_DROP_DIR no existe o no es un directorio: {self.drop_dir}"
            )

    def _latest_file(self, table: str) -> Optional[Path]:
        """Busca <TABLA>*.csv o <TABLA>*.xlsx (el más reciente gana)."""
        candidates = []
        for ext in ("csv", "xlsx", "xls"):
            candidates.extend(self.drop_dir.glob(f"{table}*.{ext}"))
            candidates.extend(self.drop_dir.glob(f"{table}*.{ext.upper()}"))
        if not candidates:
            return None
        return max(candidates, key=lambda p: p.stat().st_mtime)

    def _iter_xlsx(self, path: Path) -> Iterator[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError("Falta openpyxl: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c) if c is not None else f"col{j}"
                          for j, c in enumerate(row)]
                continue
            yield {header[j]: ("" if v is None else v)
                   for j, v in enumerate(row) if j < len(header)}
        wb.close()

    def fetch_table(self, table: str, since: Optional[str] = None,
                    since_col: Optional[str] = None) -> Iterator[dict]:
        f = self._latest_file(table)
        if not f:
            raise FileNotFoundError(
                f"No hay archivo para {table} en {self.drop_dir} "
                f"(busqué {table}*.csv, {table}*.xlsx)"
            )
        if f.suffix.lower() in (".xlsx", ".xls"):
            rows_iter = self._iter_xlsx(f)
        else:
            def _iter_csv():
                with open(f, newline="", encoding="utf-8-sig") as fh:
                    for r in self._csv.DictReader(fh):
                        yield r
            rows_iter = _iter_csv()

        for row in rows_iter:
            if since and since_col:
                val = str(row.get(since_col, ""))[:10]
                if val and val < since:
                    continue
            yield row

    def count_table(self, table: str) -> int:
        f = self._latest_file(table)
        if not f:
            return -1
        if f.suffix.lower() in (".xlsx", ".xls"):
            return sum(1 for _ in self._iter_xlsx(f))
        with open(f, newline="", encoding="utf-8-sig") as fh:
            return sum(1 for _ in fh) - 1


# ---------------------------------------------------------------------------
# Factory
# ---------------------------------------------------------------------------

def get_connection() -> SAPConnection:
    mode = (_env("SAP_MODE", "service_layer") or "service_layer").lower()
    if mode == "service_layer":
        return ServiceLayerConnection()
    if mode == "hana":
        return HanaConnection()
    if mode == "mssql":
        return MssqlConnection()
    if mode == "csv_drop":
        return CsvDropConnection()
    raise ValueError(f"SAP_MODE desconocido: {mode}")


def _cli() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--test", action="store_true",
                    help="Prueba conexión y cuenta filas de OITM.")
    args = ap.parse_args()

    if args.test:
        try:
            with get_connection() as conn:
                print(f"Conexión OK — modo: {conn.mode}")
                total = conn.count_table("OITM")
                print(f"OITM (artículos): {total} filas")
            return 0
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1
    ap.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(_cli())
